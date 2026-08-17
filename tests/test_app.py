import tempfile
import time
import unittest
from pathlib import Path

import app
import openpyxl


class AppSmokeTests(unittest.TestCase):
    def setUp(self):
        self.client = app.app.test_client()
        self.headers = {"X-Toolbox-Token": app.SESSION_TOKEN}

    def test_home_renders_ai_assistant(self):
        response = self.client.get("/")
        self.assertEqual(200, response.status_code)
        self.assertIn(b'id="p-ai"', response.data)
        self.assertIn("文本/图片 → Excel".encode(), response.data)
        self.assertIn("Excel → SQL/PDF/Word".encode(), response.data)

    def test_home_includes_responsive_and_cross_platform_ui_fixes(self):
        response = self.client.get("/")
        self.assertEqual(200, response.status_code)
        self.assertIn("Windows 10/11 与 macOS".encode(), response.data)
        self.assertIn("标题(PDF/Word)".encode(), response.data)
        self.assertIn(b"ResizeObserver", response.data)
        self.assertIn(b"#msg.warn", response.data)
        self.assertIn(b"switchPage('one')", response.data)
        self.assertIn(b"q('#im-head').checked && rows.length>0", response.data)
        self.assertIn(b"@media(max-width:360px)", response.data)
        self.assertIn(b'role="status" aria-live="polite"', response.data)
        self.assertIn(b"initAccessibility()", response.data)
        self.assertIn(b"async function requestJson", response.data)
        self.assertIn("资源管理器 / Finder".encode(), response.data)
        self.assertIn(b"x.col??x.column", response.data)
        self.assertIn(b"logs_total:j.logs_total", response.data)
        self.assertIn(b"r.warnings?.length?'warn':'ok'", response.data)

    def test_write_api_requires_session_token(self):
        response = self.client.post("/api/ai/audit", json={"sql": "SELECT 1"})
        self.assertEqual(403, response.status_code)

    def test_static_audit_route(self):
        response = self.client.post("/api/ai/audit",
                                    json={"sql": "SELECT id FROM demo LIMIT 1"},
                                    headers=self.headers)
        self.assertEqual(200, response.status_code)
        self.assertTrue(response.json["audit"]["safe"])

    def test_ai_logic_route_exists(self):
        routes = {rule.rule for rule in app.app.url_map.iter_rules()}
        self.assertIn("/api/sql/ai-logic", routes)

    def test_ai_logic_combines_static_evidence_and_model_result(self):
        class FakeAI:
            def compare_sql(self, sql_a, sql_b, evidence, **kwargs):
                self.evidence = evidence
                return {
                    "conclusion": "JOIN semantics changed",
                    "confidence": "high",
                    "evidence_basis": ["LEFT became INNER"],
                    "differences": [], "hypotheses": [],
                    "verification_steps": [], "blind_spots": [],
                }

        original = app.AI_RUNTIME
        fake = FakeAI()
        app.AI_RUNTIME = fake
        try:
            response = self.client.post("/api/sql/ai-logic", json={
                "sql_a": "SELECT a.id FROM orders a LEFT JOIN items b ON a.id=b.id",
                "sql_b": "SELECT a.id FROM orders a INNER JOIN items b ON a.id=b.id",
                "context": "retain all orders",
            }, headers=self.headers)
        finally:
            app.AI_RUNTIME = original
        self.assertEqual(200, response.status_code)
        self.assertTrue(response.json["ok"])
        self.assertEqual("high", response.json["confidence"])
        self.assertIn("joins", fake.evidence)
        self.assertIn("static", response.json)

    def test_completed_job_is_consumed_after_poll(self):
        jid = "completed-job-test"
        with app.JOBS_LOCK:
            app.JOBS[jid] = {
                "done": True,
                "msg": "完成",
                "result": {"value": 42},
                "error": None,
                "finished_at": time.time(),
            }
        response = self.client.get(f"/api/job/{jid}")
        self.assertTrue(response.json["ok"])
        self.assertEqual(42, response.json["value"])
        with app.JOBS_LOCK:
            self.assertNotIn(jid, app.JOBS)
        self.assertFalse(self.client.get(f"/api/job/{jid}").json["ok"])

    def test_clean_comparison_can_update_archive_scope(self):
        original_data_dir = app.CLIENT.data_dir
        with tempfile.TemporaryDirectory() as tmp:
            app.CLIENT.data_dir = Path(tmp)
            try:
                response = self.client.post("/api/sql/accumulate-diffs", json={
                    "name": "clean-check",
                    "diffs": [],
                    "time_slice": "复核轮次",
                    "pks_in_scope": ["1"],
                }, headers=self.headers)
            finally:
                app.CLIENT.data_dir = original_data_dir
        self.assertTrue(response.json["ok"])
        self.assertEqual(0, response.json["total"])

    def test_text_export_uses_all_rows_not_preview_slice(self):
        text = "编号\t名称\n" + "\n".join(
            f"{i:04d}\t第{i}行" for i in range(1, 506)
        )
        original_data_dir = app.CLIENT.data_dir
        with tempfile.TemporaryDirectory() as tmp:
            app.CLIENT.data_dir = Path(tmp)
            try:
                preview = self.client.post("/api/text2rows", json={
                    "text": text, "delim": "\t",
                }, headers=self.headers)
                response = self.client.post("/api/text2xlsx", json={
                    "text": text, "delim": "\t", "head": True,
                    "name": "完整文本导出",
                }, headers=self.headers)
                output = Path(tmp) / response.json["file"]
            finally:
                app.CLIENT.data_dir = original_data_dir

            self.assertTrue(preview.json["ok"])
            self.assertEqual(506, preview.json["total"])
            self.assertEqual(500, len(preview.json["rows"]))
            self.assertTrue(response.json["ok"])
            self.assertEqual(506, response.json["total"])
            self.assertEqual(505, response.json["rows"])
            wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
            try:
                ws = wb["数据"]
                self.assertEqual(506, ws.max_row)
                self.assertEqual("0505", ws["A506"].value)
            finally:
                wb.close()

    def test_empty_text_export_returns_a_clear_error(self):
        response = self.client.post(
            "/api/text2xlsx", json={"text": "  \n\t", "head": True},
            headers=self.headers,
        )
        self.assertFalse(response.json["ok"])
        self.assertIn("没有可导出", response.json["msg"])

    def test_sql_log_export_explicitly_marks_preview_truncation(self):
        original_data_dir = app.CLIENT.data_dir
        with tempfile.TemporaryDirectory() as tmp:
            app.CLIENT.data_dir = Path(tmp)
            try:
                response = self.client.post("/api/sql/export-log", json={
                    "name": "截断标注",
                    "logs": [{"level": "high", "key": "1", "column": "x",
                              "reason": "测试", "detail": "测试"}],
                    "logs_total": 3,
                    "diffs": [{"key": "1", "col": "x", "a": "A", "b": "B"}],
                    "diffs_total": 2,
                    "stats": {"keys": ["id"], "diff_cells": 2},
                    "verdict": {"text": "有差异", "next": "复核"},
                    "executed": {"sql_a": "SELECT 1", "sql_b": "SELECT 1"},
                }, headers=self.headers)
                output = Path(tmp) / response.json["file"]
            finally:
                app.CLIENT.data_dir = original_data_dir
            self.assertTrue(response.json["ok"])
            self.assertTrue(response.json["truncated"])
            self.assertIn("安全上限", response.json["warning"])
            wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
            try:
                summary = dict(wb["结论"].iter_rows(min_row=2, values_only=True))
            finally:
                wb.close()
            self.assertEqual("1/2", summary["差异明细（导出/总数）"])
            self.assertEqual("1/3", summary["排查日志（导出/总数）"])

    def test_files_lists_only_supported_artifact_types(self):
        original_data_dir = app.CLIENT.data_dir
        with tempfile.TemporaryDirectory() as tmp:
            app.CLIENT.data_dir = Path(tmp)
            try:
                supported = {
                    "report.xlsx", "rows.csv", "insert.sql",
                    "document.docx", "print.pdf",
                }
                for name in supported | {"notes.txt", "program.exe", ".hidden.xlsx"}:
                    (Path(tmp) / name).write_text("fixture", encoding="utf-8")
                response = self.client.get("/api/files")
            finally:
                app.CLIENT.data_dir = original_data_dir

        self.assertTrue(response.json["ok"])
        names = {item["name"] for item in response.json["items"]}
        self.assertEqual(supported, names)

    def test_download_only_serves_supported_visible_artifacts(self):
        original_data_dir = app.CLIENT.data_dir
        with tempfile.TemporaryDirectory() as tmp:
            app.CLIENT.data_dir = Path(tmp)
            try:
                for name in ("report.xlsx", "notes.txt", "program.exe",
                             ".hidden.xlsx"):
                    (Path(tmp) / name).write_text("fixture", encoding="utf-8")
                response = self.client.get("/download/report.xlsx")
                self.assertEqual(200, response.status_code)
                response.close()
                self.assertEqual(404, self.client.get("/download/notes.txt").status_code)
                self.assertEqual(404, self.client.get("/download/program.exe").status_code)
                self.assertEqual(404, self.client.get("/download/.hidden.xlsx").status_code)
                self.assertEqual(404, self.client.get("/download/missing.pdf").status_code)
            finally:
                app.CLIENT.data_dir = original_data_dir

    def test_xdiff_tokens_cannot_escape_upload_directory(self):
        original_upload_dir = app.UPLOAD_DIR
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            uploads = root / "uploads"
            uploads.mkdir()
            (root / "outside.xlsx").write_bytes(b"not a workbook")
            (uploads / "inside.exe").write_bytes(b"not a table")
            app.UPLOAD_DIR = uploads
            try:
                payload = {
                    "token_a": "../outside.xlsx",
                    "token_b": "../outside.xlsx",
                    "keys": ["id"],
                }
                compare = self.client.post(
                    "/api/xdiff/compare", json=payload, headers=self.headers)
                export = self.client.post(
                    "/api/xdiff/export", json=payload, headers=self.headers)
                wrong_type = self.client.post(
                    "/api/xdiff/compare",
                    json={**payload, "token_a": "inside.exe",
                          "token_b": "inside.exe"},
                    headers=self.headers,
                )
            finally:
                app.UPLOAD_DIR = original_upload_dir

        self.assertFalse(compare.json["ok"])
        self.assertFalse(export.json["ok"])
        self.assertFalse(wrong_type.json["ok"])
        self.assertIn("失效", compare.json["msg"])

    def test_upload_routes_return_clear_validation_errors(self):
        image = self.client.post(
            "/api/image2rows", data={}, headers=self.headers)
        xdiff = self.client.post(
            "/api/xdiff/upload", data={}, headers=self.headers)
        self.assertFalse(image.json["ok"])
        self.assertIn("请选择图片", image.json["msg"])
        self.assertFalse(xdiff.json["ok"])
        self.assertIn("A、B", xdiff.json["msg"])


if __name__ == "__main__":
    unittest.main()
