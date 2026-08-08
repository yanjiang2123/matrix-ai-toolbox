
from __future__ import annotations
import base64
import csv
import hashlib
import json
import os
import re
import subprocess
import sys
import time
from pathlib import Path

from ai_tools import audit_readonly_sql
def res_path(rel: str) -> Path:
    """资源文件路径（打包后落在 sys._MEIPASS 临时目录）"""
    base = getattr(sys, "_MEIPASS", None)
    return Path(base) / rel if base else APP_DIR / rel
APP_DIR = Path(__file__).resolve().parent
IS_FROZEN = getattr(sys, "frozen", False)

# 可写缓存目录：打包后不能写进 .app 内部，统一放用户目录
CACHE_DIR = Path.home() / ".matrix_toolbox" / "cache"


class ConfigError(RuntimeError):
    """配置缺失或路径不存在"""


def load_config() -> dict:
    """优先读取本机私有配置；没有时使用仓库内的空白示例配置。"""
    candidates = []
    if IS_FROZEN:
        # .app/Contents/MacOS/toolbox → 向上找到 .app 所在目录
        exe_dir = Path(sys.executable).resolve().parent
        app_parent = (exe_dir.parents[2] if len(exe_dir.parents) >= 3 else exe_dir)
        candidates += [exe_dir / "config.local.json", app_parent / "config.local.json"]
    candidates += [APP_DIR / "config.local.json", APP_DIR / "config.json",
                   res_path("config.example.json"), APP_DIR / "config.example.json"]
    for p in candidates:
        if p.exists():
            cfg = json.loads(p.read_text(encoding="utf-8"))
            cfg["_config_path"] = str(p)
            return cfg
    raise ConfigError(f"找不到本机配置或 config.example.json，已尝试: {[str(c) for c in candidates]}")


# ── 内嵌 Java 查询器 ────────────────────────────────────────────
# 与 scripts/query_db.py 同源，改动两处：
#   1. 行数上限由环境变量 MATRIX_MAX_ROWS 控制（原为硬编码 500）
#   2. 固定 JSON 输出，避免调用方再做文本解析
JAVA_SRC = r"""
import java.io.*;
import java.net.URL;
import java.net.URLClassLoader;
import java.nio.charset.StandardCharsets;
import java.sql.*;
import java.util.*;

public class ToolboxQuery {
    public static void main(String[] args) {
        if (args.length < 2) {
            System.err.println("Usage: java ToolboxQuery <driverJar> <driverClass> (connection data via stdin)");
            System.exit(1);
        }
        String driverJar   = args[0];
        String driverClass = args[1];
        String jdbcUrl;
        String username;
        String password;
        String sql;
        try {
            BufferedReader input = new BufferedReader(
                    new InputStreamReader(System.in, StandardCharsets.UTF_8));
            jdbcUrl  = decodeInput(input.readLine());
            username = decodeInput(input.readLine());
            password = decodeInput(input.readLine());
            sql      = decodeInput(input.readLine());
        } catch (Exception e) {
            System.err.println("Invalid stdin payload: " + e.getMessage());
            System.exit(1);
            return;
        }
        int maxRows = 500;
        try {
            String mr = System.getenv("MATRIX_MAX_ROWS");
            if (mr != null && !mr.isEmpty()) maxRows = Integer.parseInt(mr);
        } catch (Exception ignore) {}

        File jarFile = new File(driverJar);
        if (!jarFile.exists()) {
            System.err.println("Driver jar not found: " + jarFile.getAbsolutePath());
            System.exit(2);
        }
        try {
            URL jarUrl = jarFile.toURI().toURL();
            URLClassLoader loader = new URLClassLoader(new URL[]{jarUrl},
                    ToolboxQuery.class.getClassLoader());
            Class<?> clazz = Class.forName(driverClass, true, loader);
            Driver rawDriver = (Driver) clazz.newInstance();
            DriverManager.registerDriver(new Driver() {
                public Connection connect(String u, Properties p) throws SQLException { return rawDriver.connect(u, p); }
                public boolean acceptsURL(String u) throws SQLException { return rawDriver.acceptsURL(u); }
                public DriverPropertyInfo[] getPropertyInfo(String u, Properties p) throws SQLException { return rawDriver.getPropertyInfo(u, p); }
                public int getMajorVersion() { return rawDriver.getMajorVersion(); }
                public int getMinorVersion() { return rawDriver.getMinorVersion(); }
                public boolean jdbcCompliant() { return rawDriver.jdbcCompliant(); }
                public java.util.logging.Logger getParentLogger() throws SQLFeatureNotSupportedException { return rawDriver.getParentLogger(); }
            });
            Properties props = new Properties();
            if (!username.isEmpty()) props.setProperty("user", username);
            if (!password.isEmpty()) props.setProperty("password", password);
            try (Connection conn = DriverManager.getConnection(jdbcUrl, props)) {
                try (PreparedStatement ps = conn.prepareStatement(sql);
                     ResultSet rs = ps.executeQuery()) {
                    int cols = rs.getMetaData().getColumnCount();
                    List<String> headers = new ArrayList<>();
                    for (int i = 1; i <= cols; i++) headers.add(rs.getMetaData().getColumnLabel(i));
                    List<List<Object>> rows = new ArrayList<>();
                    int rowCount = 0;
                    while (rs.next() && rowCount < maxRows) {
                        List<Object> row = new ArrayList<>();
                        for (int i = 1; i <= cols; i++) row.add(rs.getObject(i));
                        rows.add(row);
                        rowCount++;
                    }
                    System.out.println(toJson(headers, rows, rowCount >= maxRows));
                }
            }
        } catch (Exception e) {
            System.err.println("查询失败: " + e.getClass().getName() + " - " + e.getMessage());
            System.exit(10);
        }
    }

    private static String decodeInput(String line) {
        if (line == null || line.isEmpty()) return "";
        return new String(Base64.getDecoder().decode(line), StandardCharsets.UTF_8);
    }

    private static String toJson(List<String> headers, List<List<Object>> rows, boolean truncated) {
        StringBuilder sb = new StringBuilder();
        sb.append("{\"headers\":").append(jsonArr(headers)).append(",\"rows\":[");
        for (int i = 0; i < rows.size(); i++) {
            if (i > 0) sb.append(',');
            sb.append(jsonArr(rows.get(i)));
        }
        sb.append("],\"truncated\":").append(truncated).append('}');
        return sb.toString();
    }

    private static String jsonArr(List<?> list) {
        StringBuilder sb = new StringBuilder("[");
        for (int i = 0; i < list.size(); i++) {
            if (i > 0) sb.append(',');
            sb.append(jsonVal(list.get(i)));
        }
        return sb.append(']').toString();
    }

    private static String jsonVal(Object v) {
        if (v == null) return "null";
        if (v instanceof Number || v instanceof Boolean) return String.valueOf(v);
        return "\"" + escapeJson(String.valueOf(v)) + "\"";
    }

    private static String escapeJson(String s) {
        StringBuilder sb = new StringBuilder();
        for (char c : s.toCharArray()) {
            switch (c) {
                case '"':  sb.append("\\\""); break;
                case '\\': sb.append("\\\\"); break;
                case '\b': sb.append("\\b");  break;
                case '\f': sb.append("\\f");  break;
                case '\n': sb.append("\\n");  break;
                case '\r': sb.append("\\r");  break;
                case '\t': sb.append("\\t");  break;
                default:
                    if (c < 0x20) sb.append(String.format("\\u%04x", (int) c));
                    else sb.append(c);
            }
        }
        return sb.toString();
    }
}
""".strip()

READONLY_HEADS = ("select", "show", "desc", "describe", "explain", "with")


# ── 连库引擎 ────────────────────────────────────────────────────


def clean_subprocess_env() -> dict:
    """构造干净的子进程环境。

    PyInstaller 打包后会把 DYLD_LIBRARY_PATH 等指向自己的临时解包目录，
    子进程（这里是 java）继承后会加载错误的动态库而启动失败。
    PyInstaller 把被覆盖前的原值保存在 `{VAR}_ORIG`，这里按官方建议还原。
    """
    env = dict(os.environ)
    for var in ("DYLD_LIBRARY_PATH", "DYLD_FRAMEWORK_PATH",
                "LD_LIBRARY_PATH", "LIBPATH"):
        orig = env.pop(f"{var}_ORIG", None)
        if orig:
            env[var] = orig
        else:
            env.pop(var, None)
    return env


class MatrixClient:
    """JDBC 只读查询客户端。一次构造，多次查询，复用编译好的 class。

    连接参数可以来自本机私有配置，也可以由本机前端在运行时注入。
    前端注入的密码只保存在当前 Python 进程内存中，不写回磁盘。
    """

    def __init__(self, cfg: dict):
        self.cfg = cfg
        paths = cfg.get("paths") or {}
        ws = Path(paths.get("workspace") or ".").expanduser()
        self.workspace = ws
        self.jdk_bin = self._resolve(paths.get("jdk_bin") or "", ws)
        self.jar = self._resolve(paths.get("driver_jar") or "", ws)
        self.data_dir = self._resolve(paths.get("data_dir", "data"), ws)
        m = cfg.get("matrix") or {}
        self.driver_class = str(m.get("driver_class") or "").strip()
        self.url_template = str(m.get("url_template") or "").strip()
        self.username = str(m.get("username") or "")
        self.password = str(m.get("password") or "")
        self.clusters = m.get("clusters") or {}
        self.timeout = int(m.get("sql_timeout_seconds", 180))
        self._class_dir: Path | None = None

    @staticmethod
    def _resolve(value: str, ws: Path) -> Path:
        p = Path(value).expanduser()
        return p if p.is_absolute() else ws / p

    # ── 环境自检 ──
    def check_env(self) -> list[str]:
        problems = []
        if not self.url_template:
            problems.append("尚未配置 JDBC URL；请点击页面右上角「数据库连接」")
        if not self.driver_class:
            problems.append("尚未配置 JDBC 驱动类")
        if not self.clusters:
            problems.append("尚未配置连接名称/集群参数")
        if not self._java_tool("java").exists():
            problems.append(f"缺少 java: {self.jdk_bin}（请在数据库连接面板设置 JDK bin）")
        if not self._java_tool("javac").exists():
            problems.append(f"缺少 javac: {self.jdk_bin}（需要完整 JDK，不能只有 JRE）")
        if not self.jar.is_file():
            problems.append(f"缺少 JDBC 驱动 JAR: {self.jar}")
        if not self.data_dir.exists():
            problems.append(f"产物目录不存在: {self.data_dir}（将尝试自动创建）")
        return problems

    def _java_tool(self, name: str) -> Path:
        """兼容 macOS/Linux 的 java 与 Windows 的 java.exe。"""
        plain = self.jdk_bin / name
        win = self.jdk_bin / f"{name}.exe"
        return win if os.name == "nt" and win.exists() else plain

    def configured(self) -> bool:
        return bool(self.url_template and self.driver_class and self.clusters)

    def public_state(self) -> dict:
        """返回给前端的非敏感连接状态；绝不返回账号、密码或 JDBC URL。"""
        return {
            "configured": self.configured(),
            "clusters": self.cluster_labels(),
            "default_cluster": self.default_cluster_label(),
            "driver_ready": self.jar.is_file(),
            "jdk_ready": self._java_tool("java").is_file()
                         and self._java_tool("javac").is_file(),
        }

    def cluster_labels(self) -> list[str]:
        return [v["label"] for v in self.clusters.values()]

    def cluster_by_label(self, label: str) -> str:
        for v in self.clusters.values():
            if v["label"] == label:
                return v["name"]
        raise ConfigError(f"未知集群: {label}")

    def default_cluster_label(self) -> str:
        for v in self.clusters.values():
            if v.get("default"):
                return v["label"]
        return next(iter(self.clusters.values()))["label"] if self.clusters else ""

    # ── Java 编译（带缓存，源码变了才重编）──
    def _ensure_compiled(self) -> Path:
        if self._class_dir and (self._class_dir / "ToolboxQuery.class").exists():
            return self._class_dir
        tag = hashlib.sha256(JAVA_SRC.encode()).hexdigest()[:16]
        class_dir = CACHE_DIR / tag
        marker = class_dir / "ToolboxQuery.class"
        if marker.exists():
            self._class_dir = class_dir
            return class_dir
        class_dir.mkdir(parents=True, exist_ok=True)
        src = class_dir / "ToolboxQuery.java"
        src.write_text(JAVA_SRC, encoding="utf-8")
        r = subprocess.run([str(self._java_tool("javac")), "-encoding", "UTF-8", str(src)],
                           cwd=str(class_dir), encoding="utf-8", errors="replace",
                           capture_output=True, timeout=60,
                           env=clean_subprocess_env())
        if r.returncode != 0:
            raise RuntimeError("javac 编译失败:\n" + (r.stderr or "")[:800])
        self._class_dir = class_dir
        return class_dir

    def jdbc_url(self, cluster_name: str, db: str) -> str:
        if not self.url_template:
            raise ConfigError("尚未配置 JDBC URL")
        try:
            return self.url_template.format(cluster=cluster_name, db=db)
        except KeyError as e:
            raise ConfigError(f"JDBC URL 模板包含未知占位符: {e}") from e

    def query(self, sql: str, cluster_label: str, db: str,
              max_rows: int = 500, timeout: int | None = None) -> dict:
        """执行 SQL，返回 {headers, rows, truncated}；失败抛 RuntimeError"""
        sql = sql.strip().rstrip(";")
        if not sql:
            raise RuntimeError("SQL 不能为空")
        audit = audit_readonly_sql(sql)
        if not audit["safe"]:
            reasons = "；".join(x["message"] for x in audit["issues"]
                               if x["level"] == "blocked")
            raise PermissionError("只读安全检查已拦截：" + reasons)
        if not self.configured():
            raise ConfigError("数据库连接尚未配置，请先在页面右上角设置连接")
        class_dir = self._ensure_compiled()
        env = {
            **clean_subprocess_env(),
            "PATH": f"{self.jdk_bin}{os.pathsep}{os.environ.get('PATH', '')}",
            "MATRIX_MAX_ROWS": str(max_rows),
            "LANG": "zh_CN.UTF-8",
            "LC_ALL": "zh_CN.UTF-8",
        }
        cmd = [str(self._java_tool("java")), "-Dfile.encoding=UTF-8",
               "-cp", f"{class_dir}{os.pathsep}{self.jar}",
               "ToolboxQuery", str(self.jar), self.driver_class]
        values = (self.jdbc_url(self.cluster_by_label(cluster_label), db),
                  self.username, self.password, sql)
        input_payload = "\n".join(
            base64.b64encode(v.encode("utf-8")).decode("ascii") for v in values
        ) + "\n"
        try:
            r = subprocess.run(cmd, input=input_payload,
                               encoding="utf-8", errors="replace", capture_output=True,
                               timeout=timeout or self.timeout, env=env)
        except subprocess.TimeoutExpired:
            raise RuntimeError(f"查询超时（{timeout or self.timeout}s），请缩小范围或加 LIMIT")
        payload = next((l for l in r.stdout.splitlines() if l.startswith("{")), None)
        if payload is None:
            # 把退出码和两路输出都带出来，"无输出"这种提示等于没说
            detail = _pick_error(r.stderr) or _pick_error(r.stdout)
            raise RuntimeError(
                detail or f"java 无有效输出（exit={r.returncode}）\n"
                          f"stdout: {r.stdout[:300] or '(空)'}\n"
                          f"stderr: {r.stderr[:300] or '(空)'}")
        return json.loads(payload)

    @staticmethod
    def is_readonly(sql: str) -> bool:
        # 不能只看第一个单词：WITH 后可嵌写操作，多语句也可能用 SELECT 开头绕过。
        # 静态检查是第一道防误操作保护；数据库账号仍必须保持只读最小权限。
        return bool(audit_readonly_sql(sql).get("safe"))


def _pick_error(text: str | None) -> str:
    """从 stderr 里挑出人类可读的关键错误行，滤掉 hutool DEBUG 噪音"""
    if not text:
        return ""
    # 噪音特征：hutool 日志、DEBUG 前缀
    lines = [l.strip() for l in text.splitlines()
             if l.strip() and "[DEBUG]" not in l and "cn.hutool" not in l]
    key = [l for l in lines
           if any(k in l for k in ("查询失败", "Exception", "Error", "错误", "失败", "not found"))]
    return explain_db_error("\n".join(key[:5]) if key else "\n".join(lines[-5:]))


# 数据库原话 → 加一句「这是什么意思、该怎么改」。
# StarRocks 的报错只说结论不说原因，直接抛给用户等于让他自己再查一遍。
ERROR_HINTS = [
    (r"Column '([^']+)' cannot be resolved",
     "字段「{0}」在当前查询层次里不存在。\n"
     "最常见的是「外层引用了内层字段」：子查询里有这个字段，但没 select 出来，"
     "外层就看不到它（StarRocks 不会自动往里找）。\n"
     "改法：在子查询的 SELECT 里把该字段原样带出来并逐层向上传递，"
     "或者外层改用子查询实际输出的列名。"),
    (r"Unknown table '([^']+)'|Table '([^']+)' doesn't exist|Unknown database",
     "库或表名对不上。检查：① 库名前缀是否漏了；② 是否连到了另一个集群"
     "（不同环境的表名可能不一样，如 source_xxx / target_source_xxx）。"),
    (r"Memory of .* exceed limit|exceeded memory limit|Memory limit exceeded",
     "单查询内存超限。先缩小时间范围或加过滤条件；"
     "确实要跑大范围就在同一会话里先 SET exec_mem_limit，再跑查询。"),
    (r"Getting analyzing error.*Unknown column",
     "字段名拼错或不在该表。用 DESC 库.表 核对一下真实字段名。"),
]


def explain_db_error(msg: str) -> str:
    """给数据库原始报错补一段中文解释与改法"""
    if not msg:
        return msg
    for pat, tip in ERROR_HINTS:
        m = re.search(pat, msg, re.I)
        if m:
            groups = [g for g in m.groups() if g] or [""]
            return f"{msg}\n\n—— 怎么改 ——\n{tip.format(*groups)}"
    return msg


# ── 业务能力：行数统计 / 双库对比 ────────────────────────────────

# information_schema.table_rows 对视图/新建表可能为 NULL，调用方需感知
SQL_TABLE_ROWS = (
    "SELECT table_name, table_rows, table_type FROM information_schema.tables "
    "WHERE table_schema = '{db}' ORDER BY table_name"
)


def fetch_table_rows(client: MatrixClient, cluster_label: str, db: str,
                     max_rows: int = 5000) -> list[dict]:
    """取一个库的全表行数。返回 [{name, rows, type}]，rows 为 None 表示统计信息缺失"""
    data = client.query(SQL_TABLE_ROWS.format(db=db), cluster_label, db,
                        max_rows=max_rows)
    out = []
    for r in data["rows"]:
        name = r[0]
        raw = r[1]
        ttype = r[2] if len(r) > 2 else ""
        out.append({"name": name,
                    "rows": None if raw is None else int(raw),
                    "type": ttype or ""})
    return out


def count_one(client: MatrixClient, cluster_label: str, db: str, table: str) -> int:
    """对单表 COUNT(*)，用于 table_rows 缺失时兜底"""
    data = client.query(f"SELECT COUNT(*) FROM `{db}`.`{table}`",
                        cluster_label, db, max_rows=1)
    return int(data["rows"][0][0])


def compare_two_dbs(a: list[dict], b: list[dict], mapper=None, b_filter=None) -> dict:
    """对比两个库的表清单与行数

    mapper 把 A 侧表名换算成 B 侧表名，用于源端→目标端这种「同一张表两边不同名」的场景；
    不传就按同名对比。
    b_filter 把 B 侧收窄到本次对比的范围内：拿源端的一个模块库去比整个目标库时，
    其他模块的目标表不在本次范围内，全丢进「仅B有」只会淹掉真正的异常。

    返回 only_a / only_b / diff（行数不同）/ same，外加 pairs——
    一张把两侧表名并排放的对比表，界面直接铺开就能看。
    """
    fn = mapper or (lambda n: n)
    ma = {t["name"]: t for t in a}
    mb = {t["name"]: t for t in b}
    out_of_scope = 0
    if b_filter:
        keep = {k: v for k, v in mb.items() if b_filter(k)}
        out_of_scope = len(mb) - len(keep)
        mb = keep
    used_b: set[str] = set()
    only_a, only_b, diff, same, pairs = [], [], [], [], []
    for name in sorted(ma):
        ta = ma[name]
        want = fn(name)
        tb = mb.get(want)
        if tb is None:
            only_a.append({**ta, "name_b": want})
            pairs.append(_pair(name, want, ta["rows"], None, ta["type"], "仅A有"))
            continue
        used_b.add(want)
        ra, rb = ta["rows"], tb["rows"]
        item = {"name": name, "name_b": want, "rows_a": ra, "rows_b": rb,
                "delta": (None if ra is None or rb is None else rb - ra),
                "type": ta["type"]}
        if ra is None or rb is None:
            status = "行数未知"
            diff.append(item)
        elif ra == rb:
            status = "一致"
            same.append(item)
        else:
            status = "行数不同"
            diff.append(item)
        pairs.append(_pair(name, want, ra, rb, ta["type"], status))
    for name in sorted(mb):
        if name in used_b:
            continue
        tb = mb[name]
        only_b.append({**tb, "name_b": name})
        pairs.append(_pair("", name, None, tb["rows"], tb["type"], "仅B有"))
    key = lambda x: x["name"]
    return {"only_a": sorted(only_a, key=key), "only_b": sorted(only_b, key=key),
            "diff": sorted(diff, key=key), "same": sorted(same, key=key),
            "pairs": pairs, "out_of_scope": out_of_scope}


def _pair(name_a: str, name_b: str, rows_a, rows_b, ttype: str, status: str) -> dict:
    return {"name_a": name_a, "name_b": name_b, "rows_a": rows_a, "rows_b": rows_b,
            "delta": (None if rows_a is None or rows_b is None else rows_b - rows_a),
            "type": ttype, "status": status}


# ── 表名映射：源库→目标库、不同环境改名后的对比 ─────────────────
# 可选约定：目标表名 = "target_" + 源库名 + "_" + 源表名。

MAP_MODES = {
    "same": "同名对比（两边表名一致）",
    "source_to_target": "源库 → 目标库：前缀_<源库名>_<表名>",
    "target_to_source": "目标库 → 源库：去掉目标前缀",
    "affix": "前后缀增删（例如备份表的 _old）",
    "regex": "正则替换（自己写规则）",
    "pairs": "手工指定表名对（一行一对）",
    "saved": "用已存的对应关系（连库发现后存下来的）",
}

# 连库发现出来的表名对应存到本机可写目录
PAIRS_FILE = "table_pairs.json"


def pairs_path(base_dir) -> Path:
    return Path(base_dir) / PAIRS_FILE


def load_pairs(base_dir) -> dict:
    """读已存的表名对应。文件不在或坏了都当空的，不影响其他功能"""
    p = pairs_path(base_dir)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8")) or {}
    except (ValueError, OSError):
        return {}


def pairs_key(cluster_a: str, db_a: str, cluster_b: str, db_b: str) -> str:
    """一组对应关系的归属：换了集群或库就是另一组，不能混用"""
    return f"{cluster_a}|{db_a} → {cluster_b}|{db_b}"


def discover_pairs(client: MatrixClient, cluster_a: str, db_a: str,
                   cluster_b: str, db_b: str, rule: dict | None = None,
                   progress=None) -> dict:
    """连库把两边表清单都拉下来，按规则算出真实存在的表名对应。

    规则只是「应该叫什么」，两边表清单才说明「实际有没有」。
    算出来的对应存盘复用，下次直接选「用已存的对应关系」，不必每次重配规则。
    """
    say = progress or (lambda m: None)
    say(f"读 A 侧 {db_a} 表清单…")
    names_a = [t["name"] for t in fetch_table_rows(client, cluster_a, db_a,
                                                   max_rows=5000)]
    say(f"A 侧 {len(names_a)} 张表，读 B 侧 {db_b} 表清单…")
    names_b = [t["name"] for t in fetch_table_rows(client, cluster_b, db_b,
                                                   max_rows=5000)]
    mapper, desc = build_mapper(rule, db_a, db_b)
    lower_b = {n.lower(): n for n in names_b}
    pairs, unmatched = {}, []
    for a in names_a:
        want = mapper(a)
        hit = lower_b.get(want.lower())
        if hit:
            pairs[a] = hit
        else:
            unmatched.append({"a": a, "want": want})
    say(f"配上 {len(pairs)} 对，B 侧没有对应表的 {len(unmatched)} 张")
    return {"pairs": pairs, "unmatched": unmatched[:200],
            "unmatched_total": len(unmatched), "rule_desc": desc,
            "count_a": len(names_a), "count_b": len(names_b),
            "key": pairs_key(cluster_a, db_a, cluster_b, db_b)}


def save_pairs(base_dir, key: str, pairs: dict, meta: dict | None = None) -> dict:
    """把一组表名对应写进本地文件。同一个 key 直接覆盖（重新发现就是要更新）"""
    store = load_pairs(base_dir)
    store[key] = {"pairs": pairs, "count": len(pairs),
                  "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                  **(meta or {})}
    p = pairs_path(base_dir)
    p.write_text(json.dumps(store, ensure_ascii=False, indent=2),
                 encoding="utf-8")
    return {"file": str(p), "key": key, "count": len(pairs),
            "groups": len(store)}


SYS_DBS = {"information_schema", "_statistics_", "sys", "performance_schema",
           "mysql", "__internal_schema"}


def list_databases(client: MatrixClient, cluster_label: str) -> list[str]:
    """一个集群下的库名，去掉系统库"""
    data = client.query("SHOW DATABASES", cluster_label, "", max_rows=2000)
    return [r[0] for r in data["rows"] if r[0] not in SYS_DBS]


def list_tables(client: MatrixClient, cluster_label: str, db: str) -> list[str]:
    """一个库的表名"""
    data = client.query(
        f"SELECT table_name FROM information_schema.tables "
        f"WHERE table_schema = '{db}'", cluster_label, db, max_rows=20000)
    return [r[0] for r in data["rows"]]


def split_target_name(table: str, source_dbs: list[str],
                      prefix: str = "target_") -> tuple:
    """把目标表名拆回 (源库名, 源表名)，拆不出返回 ("", "")。

    必须按库名从长到短试：`source_business_long` 与 `source_business` 若都存在，
    短的先匹配会把库名切错、表名多带一截。
    """
    if not table.startswith(prefix):
        return "", ""
    body = table[len(prefix):]
    for source_db in sorted(source_dbs, key=len, reverse=True):
        if body.startswith(source_db + "_"):
            return source_db, body[len(source_db) + 1:]
    return "", ""


def discovered_map_key(source_cluster: str, target_cluster: str,
                       target_db: str) -> str:
    """整套源库→目标库对应关系的归属（跨全部源库，与单库映射分开存）"""
    return f"[全库] {source_cluster} → {target_cluster}|{target_db}"


def discover_target_map(client: MatrixClient, source_cluster: str,
                        target_cluster: str, target_db: str = "target",
                        prefix: str = "target_", progress=None) -> dict:
    """一次算出全部源库到目标库的表对应关系。

    当源系统按模块拆成多个库时，可以反过来利用目标表的命名约定：
    若目标表名为 `target_<源库名>_<源表名>`，拿源库名清单去拆它，
    一趟就能把所有库的对应关系全拆出来。
    拆不出库名的表单独列出，不直接判定为缺失。
    """
    say = progress or (lambda m: None)
    say("读取源端连接的库名清单…")
    source_dbs = list_databases(client, source_cluster)
    say(f"源端有 {len(source_dbs)} 个库，读取目标库 {target_db} 的表清单…")
    target_tables = list_tables(client, target_cluster, target_db)
    say(f"{target_db} 有 {len(target_tables)} 张表，按库名前缀反推…")
    by_db: dict[str, list] = {}
    pairs: dict[str, str] = {}
    unknown, unknown_prefix = [], {}
    for table in target_tables:
        source_db, source_table = split_target_name(table, source_dbs, prefix)
        if not source_db:
            unknown.append(table)
            # 猜一下它的前缀是什么，便于判断是哪个源系统
            guess = (table[len(prefix):].split("_")[0]
                     if table.startswith(prefix) else "(无目标前缀)")
            unknown_prefix[guess] = unknown_prefix.get(guess, 0) + 1
            continue
        by_db.setdefault(source_db, []).append(
            {"source_table": source_table, "target_table": table})
        pairs[f"{source_db}.{source_table}"] = table
    for v in by_db.values():
        v.sort(key=lambda x: x["source_table"])
    say(f"识别出 {len(pairs)} 张、涉及 {len(by_db)} 个源库；"
        f"另有 {len(unknown)} 张无法按当前规则归类")
    return {"by_db": by_db, "pairs": pairs,
            "unknown": sorted(unknown)[:400], "unknown_total": len(unknown),
            "unknown_prefix": dict(sorted(unknown_prefix.items(),
                                          key=lambda x: -x[1])),
            "source_dbs": source_dbs, "target_db": target_db, "prefix": prefix,
            "target_total": len(target_tables), "matched": len(pairs),
            "db_count": len(by_db),
            "key": discovered_map_key(source_cluster, target_cluster, target_db)}


def save_discovered_map(base_dir, res: dict) -> dict:
    """把全库对应关系落盘。按源库分组存，库数据对比时按库名取用。"""
    store = load_pairs(base_dir)
    store[res["key"]] = {
        "kind": "discovered_map", "target_db": res["target_db"],
        "prefix": res["prefix"],
        "by_db": {k: {i["source_table"]: i["target_table"] for i in v}
                  for k, v in res["by_db"].items()},
        "count": res["matched"], "db_count": res["db_count"],
        "unknown_total": res["unknown_total"],
        "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    p = pairs_path(base_dir)
    p.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"file": str(p), "key": res["key"], "count": res["matched"],
            "db_count": res["db_count"], "groups": len(store)}


def saved_pairs_for(base_dir, cluster_a: str, db_a: str,
                    cluster_b: str, db_b: str) -> tuple[dict, str]:
    """取这次对比该用哪套表名对应，返回 (表名对, 来源说明)。

    先找「这一对库」单独存过的；没有再从整套源端→目标端全库对应里按库名切一份出来。
    """
    store = load_pairs(base_dir)
    exact = store.get(pairs_key(cluster_a, db_a, cluster_b, db_b)) or {}
    if exact.get("pairs"):
        return exact["pairs"], f"已存的这对库的对应（{exact.get('saved_at', '')}）"
    for key, item in store.items():
        if item.get("kind") != "discovered_map":
            continue
        if not key.endswith(f"{cluster_b}|{db_b}") or f"] {cluster_a} " not in key:
            continue
        sub = (item.get("by_db") or {}).get(db_a)
        if sub:
            return sub, (f"从全库对应里取 {db_a} 这一段"
                         f"（共 {len(sub)} 张，{item.get('saved_at', '')}）")
    return {}, ""


def edit_pairs(base_dir, key: str, add: dict | None = None,
               drop: list[str] | None = None, replace: bool = False,
               meta: dict | None = None) -> dict:
    """手工维护一组表名对应：合并新增、删掉指定项，或整组替换。

    自动发现只能覆盖有规律的那部分，规则不规则的（改过名、拆表、并表）
    只能手工指定，所以这套对应必须能改、能删、能一直攒。
    """
    store = load_pairs(base_dir)
    item = store.get(key) or {}
    cur = {} if replace else dict(item.get("pairs") or {})
    for a in (drop or []):
        cur.pop(a, None)
    cur.update(add or {})
    if not cur:
        store.pop(key, None)
    else:
        store[key] = {**item, "pairs": cur, "count": len(cur),
                      "kind": item.get("kind") or "manual",
                      "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                      **(meta or {})}
    p = pairs_path(base_dir)
    p.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"key": key, "count": len(cur), "groups": len(store),
            "file": str(p), "removed": not cur}


def pairs_groups(base_dir) -> list[dict]:
    """已存的所有对应关系分组概览"""
    out = []
    for key, item in load_pairs(base_dir).items():
        kind = item.get("kind") or "manual"
        out.append({"key": key, "kind": kind,
                    "count": item.get("count", 0),
                    "db_count": item.get("db_count", 0),
                    "saved_at": item.get("saved_at", ""),
                    "rule_desc": item.get("rule_desc", ""),
                    "dbs": sorted((item.get("by_db") or {}).keys())})
    out.sort(key=lambda x: (x["kind"] != "discovered_map", x["key"]))
    return out


def pairs_of_group(base_dir, key: str, db: str = "", kw: str = "",
                   limit: int = 500) -> dict:
    """取一组对应关系的明细，可按源库与关键词筛。"""
    item = load_pairs(base_dir).get(key) or {}
    if item.get("kind") == "discovered_map":
        by = item.get("by_db") or {}
        dbs = [db] if db else sorted(by)
        items = [{"db": d, "a": a, "b": b}
                 for d in dbs for a, b in sorted((by.get(d) or {}).items())]
    else:
        items = [{"db": "", "a": a, "b": b}
                 for a, b in sorted((item.get("pairs") or {}).items())]
    if kw:
        low = kw.lower()
        items = [i for i in items
                 if low in i["a"].lower() or low in i["b"].lower()
                 or low in i["db"].lower()]
    return {"key": key, "kind": item.get("kind") or "manual",
            "total": len(items), "items": items[:limit],
            "dbs": sorted((item.get("by_db") or {}).keys()),
            "saved_at": item.get("saved_at", "")}


# ── 差异跟踪档案累入 ─────────────────────────────────────────────
# 多轮 SQL 对比的差异累到一份 Excel，按「主键 + 字段」去重，
# 本轮不再出现的差异自动标「已修复」，方便跨时间片排查后回归。
ARCHIVE_HEADERS = ["主键", "字段", "A 值", "B 值", "差异类型",
                   "首次发现", "最后出现", "修复状态", "差异标签"]


def accumulate_diffs(archive_path, new_diffs: list[dict],
                     time_slice: str, pks_in_scope: list[str] | None = None
                     ) -> dict:
    """把一轮 SQL 对比的差异累入差异跟踪档案。

    参数：
      new_diffs     本轮差异，每条 {pk, col, a, b, kind}
                    col 留空表示整行缺失（only_a/only_b 那种）
      time_slice    时间片标签，写入「首次发现/最后出现」列
      pks_in_scope  本轮对比涉及的主键集合；只有范围内的历史未修复条目
                    本轮没再出现才标「已修复」，范围外的不动（没复查到）

    返回：{added, fixed, kept_unfixed, updated, total, was_empty}
    """
    archive: dict[tuple[str, str], list] = {}
    was_empty = True

    if archive_path.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(archive_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
            for r in rows:
                if not r or not r[0]:
                    continue
                pk = str(r[0])
                col = str(r[1]) if r[1] is not None else ""
                archive[(pk, col)] = [
                    pk, col,
                    r[2] if len(r) > 2 else "",
                    r[3] if len(r) > 3 else "",
                    r[4] if len(r) > 4 else "",
                    r[5] if len(r) > 5 else "",
                    r[6] if len(r) > 6 else "",
                    r[7] if len(r) > 7 else "未修复",
                    r[8] if len(r) > 8 else "",
                ]
            was_empty = not archive
        except (OSError, ValueError):
            archive = {}

    added = 0
    updated = 0
    for d in new_diffs:
        pk = str(d.get("pk", "") or "")
        col = str(d.get("col", "") or "")
        a = str(d.get("a", "") if d.get("a") is not None else "")
        b = str(d.get("b", "") if d.get("b") is not None else "")
        kind = str(d.get("kind", "") or "")
        if not pk and not col:
            continue
        key = (pk, col)
        if key in archive:
            row = archive[key]
            row[2], row[3], row[4], row[6] = a, b, kind, time_slice
            if row[7] == "已修复":
                row[7] = "又出现"
            updated += 1
        else:
            archive[key] = [pk, col, a, b, kind, time_slice, time_slice,
                            "未修复", ""]
            added += 1

    scope = {str(pk) for pk in (pks_in_scope or []) if pk is not None}
    fixed = 0
    kept_unfixed = 0
    if scope:
        for (pk, col), row in archive.items():
            if row[7] != "未修复":
                continue
            if col == "":
                # 整行缺失条目：靠 pk 是否在范围内判断是否已修复
                if pk in scope:
                    row[7] = "已修复"
                    fixed += 1
                else:
                    kept_unfixed += 1
            else:
                # 字段差异条目：pk 在范围内且本轮没再出现 → 已修复
                if pk in scope:
                    row[7] = "已修复"
                    fixed += 1
                else:
                    kept_unfixed += 1

    write_xlsx(archive_path, {"差异跟踪": (ARCHIVE_HEADERS, list(archive.values()))})
    return {"added": added, "fixed": fixed,
            "kept_unfixed": kept_unfixed, "updated": updated,
            "total": len(archive), "was_empty": was_empty}


def parse_pairs(text: str) -> dict[str, str]:
    """解析手工表名对。一行一对，支持 `A=>B`、`A→B`、`A,B`、`A\\tB` 四种写法"""
    out: dict[str, str] = {}
    for line in (text or "").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = re.split(r"\s*(?:=>|->|→|\t|,|\|)\s*", line, maxsplit=1)
        if len(parts) != 2:
            continue
        a, b = parts[0].strip().strip("`"), parts[1].strip().strip("`")
        if a and b:
            out[a.split(".")[-1]] = b.split(".")[-1]
    return out


def build_mapper(rule: dict | None, db_a: str = "", db_b: str = ""):
    """按规则生成「A 表名 → B 表名」的换算函数，返回 (fn, 人话描述)"""
    rule = rule or {}
    mode = (rule.get("mode") or "same").strip()
    if mode not in MAP_MODES:
        raise ValueError(f"未知的表名映射方式: {mode}")

    if mode == "same":
        return (lambda n: n), MAP_MODES["same"]

    if mode == "source_to_target":
        pfx = f"target_{db_a.strip()}_"
        return (lambda n: f"{pfx}{n}"), f"A 侧表名前面拼 `{pfx}`"

    if mode == "target_to_source":
        pfx = f"target_{db_b.strip()}_"
        def strip_pfx(n):
            return n[len(pfx):] if n.startswith(pfx) else n
        return strip_pfx, f"去掉 A 侧表名的 `{pfx}` 前缀"

    if mode == "affix":
        pd = (rule.get("prefix_del") or "").strip()
        pa = (rule.get("prefix_add") or "").strip()
        sd = (rule.get("suffix_del") or "").strip()
        sa = (rule.get("suffix_add") or "").strip()
        if not (pd or pa or sd or sa):
            raise ValueError("前后缀四个框都空着，等于同名对比——请至少填一个")
        def affix(n):
            if pd and n.startswith(pd):
                n = n[len(pd):]
            if sd and n.endswith(sd):
                n = n[:-len(sd)]
            return f"{pa}{n}{sa}"
        bits = [f"去前缀 `{pd}`" if pd else "", f"去后缀 `{sd}`" if sd else "",
                f"加前缀 `{pa}`" if pa else "", f"加后缀 `{sa}`" if sa else ""]
        return affix, "；".join(b for b in bits if b)

    if mode == "regex":
        pat, rep = (rule.get("pattern") or "").strip(), rule.get("replace") or ""
        if not pat:
            raise ValueError("正则模式不能为空")
        try:
            cre = re.compile(pat)
        except re.error as e:
            raise ValueError(f"正则写错了: {e}") from e
        return (lambda n: cre.sub(rep, n)), f"正则 `{pat}` → `{rep}`"

    if mode == "saved":
        pairs = rule.get("saved_pairs") or {}
        if not pairs:
            raise ValueError("这组集群+库还没存过表名对应。请切到「表对应关系」页，"
                             "在「连库发现」那一栏点 ▶ 跑一次，结果会自动存下来；"
                             "再回到这里选「用已存的对应关系」即可")
        return (lambda n: pairs.get(n, n)), f"用已存的 {len(pairs)} 对表名"

    pairs = parse_pairs(rule.get("pairs") or "")
    if not pairs:
        raise ValueError("没解析出任何表名对。每行写一对，如 "
                         "`source_table => target_source_table`")
    return (lambda n: pairs.get(n, n)), f"手工指定 {len(pairs)} 对表名"


def build_b_filter(rule: dict | None, db_a: str = "", db_b: str = ""):
    """B 侧的对比范围。

    源端→目标端是「一个模块库比整个目标库」：只有 target_<源库名>_ 开头的表才在范围内，
    其余模块的表不参与对比，否则「仅B有」会有几百条无关表。
    其他映射方式默认不收窄。
    """
    rule = rule or {}
    mode = rule.get("mode") or "same"
    if mode == "saved":
        # 用已存对应时，范围就是这套对应涉及的 B 侧表；其余是别的库/别的源系统
        want = {v.lower() for v in (rule.get("saved_pairs") or {}).values()}
        if not want:
            return None, ""
        return ((lambda n: n.lower() in want),
                f"B 侧只看对应关系里的 {len(want)} 张表")
    if mode != "source_to_target" or not db_a.strip():
        return None, ""
    pfx = f"target_{db_a.strip()}_"
    return (lambda n: n.startswith(pfx)), f"B 侧只看 `{pfx}` 开头的表"


def mapping_preview(names: list[str], rule: dict, db_a: str = "", db_b: str = "",
                    limit: int = 40) -> dict:
    """映射规则试跑：先看几个例子对不对，再决定要不要连库全量对比"""
    fn, desc = build_mapper(rule, db_a, db_b)
    items = [{"name_a": n, "name_b": fn(n), "changed": fn(n) != n}
             for n in names[:limit]]
    changed = sum(1 for i in items if i["changed"])
    return {"items": items, "desc": desc, "changed": changed,
            "total": len(names), "shown": len(items)}


# ── 单表对比：结构 + 行数 + 明细数据 ────────────────────────────

SQL_TABLE_COLUMNS = (
    "SELECT column_name, data_type, is_nullable, column_default, ordinal_position "
    "FROM information_schema.columns "
    "WHERE table_schema = '{db}' AND table_name = '{table}' "
    "ORDER BY ordinal_position"
)


def fetch_columns(client: MatrixClient, cluster_label: str, db: str,
                  table: str) -> list[dict]:
    """取一张表的字段清单，用于结构对比"""
    data = client.query(SQL_TABLE_COLUMNS.format(db=db, table=table),
                        cluster_label, db, max_rows=2000)
    return [{"name": r[0], "type": str(r[1] or ""),
             "nullable": str(r[2] or ""), "default": r[3],
             "pos": int(r[4] or 0)} for r in data["rows"]]


SQL_TABLE_FIELDS = (
    "SELECT column_name, data_type, column_key, ordinal_position "
    "FROM information_schema.columns "
    "WHERE table_schema = '{db}' AND table_name = '{table}' "
    "ORDER BY ordinal_position"
)

# 表结构里真正的时间类型。只按字段名猜会两头都错：
# 名字里没 date/time 的 datetime 列会漏，叫 xxx_date 的 varchar 会误判
TIME_TYPES = ("date", "datetime", "timestamp")
# 名字带这些词、且是时间类型的，才算业务时间；系统时间排在后面
BIZ_TIME_TAIL = ("sys_load_time", "sys_update_time", "sys_create_time",
                 "create_time", "update_time", "etl_time", "load_time")


def table_fields(client: MatrixClient, cluster_label: str, db: str,
                 table: str) -> list[dict]:
    """取一张表的字段名、类型与是否主键"""
    data = client.query(SQL_TABLE_FIELDS.format(db=db, table=table),
                        cluster_label, db, max_rows=3000)
    return [{"name": r[0], "type": str(r[1] or "").lower(),
             "key": str(r[2] or "").upper(), "pos": int(r[3] or 0)}
            for r in data["rows"]]


def table_field_hints(client: MatrixClient, cluster_label: str, db: str,
                      tables: list[str]) -> dict:
    """连库读表结构，挑出真正的时间字段与主键。

    比从 SQL 文本里猜可靠：data_type 说明它到底是不是时间，
    column_key='PRI' 直接给出建表时定的主键，不用靠字段名像不像 id 去蒙。
    """
    out, times, keys = [], [], []
    for full in tables:
        tdb, _, tname = full.rpartition(".")
        tdb = tdb or db
        item = {"table": tname, "db": tdb, "ok": False, "msg": "",
                "time_cols": [], "key_cols": [], "total": 0}
        try:
            cols = table_fields(client, cluster_label, tdb, tname)
        except Exception as e:                       # 无权限或表不存在都不该中断
            item["msg"] = _pick_error(e)
            out.append(item)
            continue
        if not cols:
            item["msg"] = f"{tdb}.{tname} 在表结构里查不到，确认库名与集群是否选对"
            out.append(item)
            continue
        item["ok"], item["total"] = True, len(cols)
        for c in cols:
            if any(c["type"].startswith(t) for t in TIME_TYPES):
                low = c["name"].lower()
                # sys_ 前缀是这批库里系统字段的统一约定，一并归到系统时间
                sys_time = low in BIZ_TIME_TAIL or low.startswith("sys_")
                item["time_cols"].append({"column": c["name"], "type": c["type"],
                                          "sys": sys_time})
            if c["key"] == "PRI":
                item["key_cols"].append({"column": c["name"], "type": c["type"]})
        item["time_cols"].sort(key=lambda x: x["sys"])   # 业务时间排前面
        for t in item["time_cols"]:
            times.append({**t, "table": tname})
        for k in item["key_cols"]:
            keys.append({**k, "table": tname})
        out.append(item)
    return {"tables": out, "time_cols": times, "key_cols": keys}


def compare_columns(ca: list[dict], cb: list[dict]) -> dict:
    """字段级结构对比：只在A / 只在B / 类型不同 / 一致

    类型只比大类（varchar(64) 与 varchar(255) 算同类但长度不同，单独标注），
    因为跨集群同步经常只差长度，全当"类型不同"会淹掉真正的差异。
    """
    ma = {c["name"].lower(): c for c in ca}
    mb = {c["name"].lower(): c for c in cb}

    def base(t):
        return re.split(r"[(<]", t.lower(), maxsplit=1)[0].strip()

    only_a = [ma[k] for k in ma if k not in mb]
    only_b = [mb[k] for k in mb if k not in ma]
    type_diff, same = [], []
    for k in ma:
        if k not in mb:
            continue
        x, y = ma[k], mb[k]
        item = {"name": x["name"], "type_a": x["type"], "type_b": y["type"],
                "nullable_a": x["nullable"], "nullable_b": y["nullable"],
                "kind": ""}
        if base(x["type"]) != base(y["type"]):
            item["kind"] = "类型不同"
        elif x["type"] != y["type"]:
            item["kind"] = "仅长度/精度不同"
        elif x["nullable"] != y["nullable"]:
            item["kind"] = "可空性不同"
        (type_diff if item["kind"] else same).append(item)
    key = lambda x: x["name"].lower()
    return {"only_a": sorted(only_a, key=key), "only_b": sorted(only_b, key=key),
            "type_diff": sorted(type_diff, key=key), "same": sorted(same, key=key),
            "common": [ma[k]["name"] for k in ma if k in mb]}


def count_table(client: MatrixClient, cluster_label: str, db: str, table: str,
                where: str = "") -> int:
    """单表 COUNT(*)，可带 WHERE 收窄范围"""
    cond = f" WHERE {where}" if where.strip() else ""
    data = client.query(f"SELECT COUNT(*) FROM `{db}`.`{table}`{cond}",
                        cluster_label, db, max_rows=1)
    return int(data["rows"][0][0]) if data["rows"] else 0


def fetch_table_sample(client: MatrixClient, cluster_label: str, db: str,
                       table: str, cols: list[str], keys: list[str],
                       where: str = "", limit: int = 500) -> tuple[list, list]:
    """按主键排序取一段明细。

    两边都「按主键排序取前 N 行」才是可比的样本；不排序的 LIMIT 在 MPP 引擎下
    每次回来的行都可能不一样，比出来的差异全是假的。
    """
    pick = ", ".join(f"`{c}`" for c in cols) if cols else "*"
    cond = f"\nWHERE {where}" if where.strip() else ""
    order = ("\nORDER BY " + ", ".join(f"`{k}`" for k in keys)) if keys else ""
    sql = f"SELECT {pick}\nFROM `{db}`.`{table}`{cond}{order}\nLIMIT {limit}"
    data = client.query(sql, cluster_label, db, max_rows=limit)
    return data["headers"], data["rows"]


def compare_one_table(client: MatrixClient, side_a: dict, side_b: dict,
                      keys: list[str] | None = None, limit: int = 500,
                      where_a: str = "", where_b: str = "",
                      progress=None) -> dict:
    """单表对比：先比结构，再比行数，给了主键就再按主键对齐比明细。

    side_x = {"cluster": 集群label, "db": 库名, "table": 表名}
    结构先比是有讲究的：字段对不上时，明细比对必然一片红，先看结构能省一轮无效排查。
    """
    import sql_tools as st

    def note(m):
        if progress:
            progress(m)

    ta = f"{side_a['db']}.{side_a['table']}"
    tb = f"{side_b['db']}.{side_b['table']}"
    note(f"读结构：{ta}")
    ca = fetch_columns(client, side_a["cluster"], side_a["db"], side_a["table"])
    note(f"读结构：{tb}")
    cb = fetch_columns(client, side_b["cluster"], side_b["db"], side_b["table"])
    if not ca:
        raise RuntimeError(f"A 侧表不存在或没有字段：{ta}")
    if not cb:
        raise RuntimeError(f"B 侧表不存在或没有字段：{tb}")
    cols = compare_columns(ca, cb)

    note(f"数行数：{ta}")
    na = count_table(client, side_a["cluster"], side_a["db"], side_a["table"], where_a)
    note(f"数行数：{tb}")
    nb = count_table(client, side_b["cluster"], side_b["db"], side_b["table"], where_b)

    out = {
        "table": {"a": ta, "b": tb},
        "cluster": {"a": side_a["cluster"], "b": side_b["cluster"]},
        "columns": {k: v for k, v in cols.items() if k != "common"},
        "col_summary": {"a": len(ca), "b": len(cb),
                        "only_a": len(cols["only_a"]), "only_b": len(cols["only_b"]),
                        "type_diff": len(cols["type_diff"]), "same": len(cols["same"])},
        "count": {"a": na, "b": nb, "delta": nb - na},
        "data": None,
        "where": {"a": where_a, "b": where_b},
    }

    keys = [k for k in (keys or []) if k.strip()]
    if not keys:
        out["note"] = ("只比了结构和行数。填上主键字段就能继续按主键对齐比明细，"
                       "定位到底是哪几条、哪个字段不一样")
        return out

    common = {c.lower() for c in cols["common"]}
    missing = [k for k in keys if k.split(".")[-1].lower() not in common]
    if missing:
        raise RuntimeError(f"主键字段在两边不都存在：{missing}。"
                           f"两边共有字段 {len(common)} 个，请从中挑主键")
    # 只取两边都有的字段：一边缺的字段没法比，带上只会让差异列表全是噪音
    pick = [c for c in cols["common"]]
    note(f"取明细：{ta}（按主键排序前 {limit} 行）")
    ha, ra = fetch_table_sample(client, side_a["cluster"], side_a["db"],
                                side_a["table"], pick, keys, where_a, limit)
    note(f"取明细：{tb}（按主键排序前 {limit} 行）")
    hb, rb = fetch_table_sample(client, side_b["cluster"], side_b["db"],
                                side_b["table"], pick, keys, where_b, limit)
    note("按主键对齐比明细…")
    out["data"] = st.compare_details(ha, ra, hb, rb, keys)
    out["data"]["sampled"] = (na > limit or nb > limit)
    out["note"] = (f"明细按主键排序各取前 {limit} 行比对"
                   + ("；两边行数都超过取数上限，这只是抽样结论，"
                      "要比全量请加 WHERE 条件分段跑" if na > limit or nb > limit else ""))
    return out


# ── Excel 工具 ──────────────────────────────────────────────────


def _grid_size(grid: list[list]) -> tuple[int, int]:
    rows = len(grid)
    cols = max((len(r) for r in grid), default=0)
    return rows, cols


def _load_grid(path: Path, sheet: str | int | None = None,
               data_only: bool = True) -> tuple[list[list], str, list[str], bool]:
    """读出某个工作表的完整网格，返回 (网格, 表名, 全部表名, 是否走了兼容模式)。

    read_only 模式靠 worksheet XML 里的 <dimension> 判断范围，速度快。
    但程序导出的 xlsx 经常把 dimension 写错（比如整张表声明成 A1:A1），
    read_only 就只读 1 行 1 列——1.2MB 的文件能读出 0 行数据。
    所以先快速读，结果可疑再回退到完整加载（慢但准）。
    """
    from openpyxl import load_workbook

    def pick(wb):
        if isinstance(sheet, str) and sheet in wb.sheetnames:
            return wb[sheet]
        if isinstance(sheet, int) and 0 <= sheet < len(wb.worksheets):
            return wb.worksheets[sheet]
        return None

    wb = load_workbook(path, read_only=True, data_only=data_only)
    try:
        names = list(wb.sheetnames)
        ws = pick(wb) or wb.worksheets[0]
        title = ws.title
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    rows, cols = _grid_size(grid)
    # 1x1 或空表，但文件明显不止这点内容 → dimension 不可信
    suspicious = (rows <= 1 or cols <= 1) and path.stat().st_size > 8192
    if not suspicious:
        return grid, title, names, False

    wb = load_workbook(path, data_only=data_only)
    try:
        names = list(wb.sheetnames)
        ws = pick(wb) or wb.worksheets[0]
        title = ws.title
        grid2 = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    rows2, cols2 = _grid_size(grid2)
    if rows2 * max(cols2, 1) > rows * max(cols, 1):
        return grid2, title, names, True
    return grid, title, names, False


def list_sheets(path: Path) -> list[dict]:
    """列出所有工作表及其有效数据规模，供调用方选择"""
    from openpyxl import load_workbook

    out = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        quick = []
        for ws in wb.worksheets:
            filled = 0
            for row in ws.iter_rows(min_row=1,
                                    max_row=min(ws.max_row or 1, 200),
                                    values_only=True):
                filled += sum(1 for v in row
                              if v is not None and str(v).strip() != "")
            quick.append({"name": ws.title, "rows": ws.max_row or 0,
                          "cols": ws.max_column or 0, "filled": filled})
    finally:
        wb.close()

    # 所有表都被 dimension 骗成 1x1 时，回退完整加载重新统计
    if (all(s["filled"] <= 1 or s["cols"] <= 1 for s in quick)
            and path.stat().st_size > 8192):
        wb = load_workbook(path, data_only=True)
        try:
            for ws in wb.worksheets:
                filled, nrow, ncol = 0, 0, 0
                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    nrow = i
                    ncol = max(ncol, len(row))
                    if i <= 200:
                        filled += sum(1 for v in row
                                      if v is not None and str(v).strip() != "")
                out.append({"name": ws.title, "rows": nrow, "cols": ncol,
                            "filled": filled})
        finally:
            wb.close()
        return out or quick
    return quick


def _pick_sheet_index(sheets: list[dict]) -> int:
    """挑有效单元格最多的表。

    很多业务 Excel 第一页是「说明」或空白封面，固定读第一页会得到空表，
    转出来的 PDF/Word 自然也是空的。
    """
    if not sheets:
        return 0
    best = max(range(len(sheets)), key=lambda i: sheets[i]["filled"])
    return best


def read_sheet_meta(path: Path, sheet: str | int | None = None
                    ) -> tuple[list[str], list[list], dict]:
    """读取工作表，额外返回读取过程的诊断信息。

    做了四件容错，都是被真实文件坑过才加的：
    1. dimension 声明不可信时自动回退完整加载（不然 1.2MB 文件能读出 0 行）
    2. sheet=None 时自动挑数据最多的那一页，而不是死读第一页
    3. 跳过表头前的空行与跨列标题行
    4. 公式单元格在 data_only 下没缓存值会读成 None，逐格回退成公式文本

    兼容模式下完整加载很贵，所以只在必要时做：单表文件不预扫、
    空洞很少时不为了公式再加载一遍。
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()

    sheets = None
    if sheet is None and len(names) > 1:
        sheets = list_sheets(path)          # 多表才值得预扫一遍来挑页
        sheet = _pick_sheet_index(sheets)

    grid, title, names, compat = _load_grid(path, sheet, data_only=True)
    nrow, ncol = _grid_size(grid)
    if sheets is None:
        filled = sum(1 for r in grid[:200] for v in r
                     if v is not None and str(v).strip() != "")
        sheets = [{"name": n,
                   "rows": nrow if n == title else 0,
                   "cols": ncol if n == title else 0,
                   "filled": filled if n == title else 0} for n in names]

    meta = {"sheet": title, "sheets": sheets, "formula_cells": 0,
            "header_row": 1, "compat_mode": compat, "warnings": []}
    if compat:
        meta["warnings"].append(
            "这个文件的表格范围声明不准（多半是程序导出的），"
            "已改用兼容模式完整扫描，读取会慢一些但结果准确")
    if len(names) > 1:
        meta["warnings"].append(
            f"该文件有 {len(names)} 个工作表，已选中「{title}」。"
            f"如需其它表请手动指定")

    # 公式回退：只在空洞多到影响结果时才为此再加载一遍
    if grid:
        holes = [(i, j) for i, r in enumerate(grid) for j, v in enumerate(r)
                 if v is None]
        cells = max(nrow * ncol, 1)
        if holes and len(holes) > 10 and len(holes) / cells > 0.03:
            raw, _, _, _ = _load_grid(path, sheet, data_only=False)
            filled = 0
            for i, j in holes:
                if i < len(raw) and j < len(raw[i]):
                    v = raw[i][j]
                    if v is not None and str(v).strip() != "":
                        grid[i][j] = v
                        filled += 1
            if filled:
                meta["formula_cells"] = filled
                meta["warnings"].append(
                    f"有 {filled} 个单元格是公式且没有缓存结果，已显示公式原文。"
                    f"想拿到计算值请用 Excel 打开该文件另存一次")

    if not grid:
        meta["warnings"].append("这一页没有任何内容")
        return [], [], meta

    # 找表头行：第一个非空单元格数达到本页最宽一半的行
    widths = [sum(1 for v in r if v is not None and str(v).strip() != "")
              for r in grid]
    max_w = max(widths) if widths else 0
    head_idx = 0
    for i, w in enumerate(widths):
        if w >= min(2, max_w) and w >= max_w * 0.5:
            head_idx = i
            break
    else:
        head_idx = next((i for i, w in enumerate(widths) if w), 0)
    meta["header_row"] = head_idx + 1
    if head_idx:
        meta["warnings"].append(
            f"前 {head_idx} 行是空行或标题行，已跳过，把第 {head_idx + 1} 行当表头")

    head_row = grid[head_idx]
    headers = [str(h).strip() if h is not None and str(h).strip() != ""
               else f"列{i+1}" for i, h in enumerate(head_row)]
    # 去掉尾部因合并单元格产生的空列
    while headers and headers[-1].startswith("列") and all(
            (r[len(headers) - 1] if len(r) >= len(headers) else None) in (None, "")
            for r in grid[head_idx + 1:]):
        headers.pop()
    rows = [r[:len(headers)] for r in grid[head_idx + 1:]
            if any(v is not None and str(v).strip() != "" for v in r)]
    if not rows:
        meta["warnings"].append("识别到表头但没有数据行")
    return headers, rows, meta


def read_sheet(path: Path, sheet: str | int | None = None,
               keep_formula: bool = True) -> tuple[list[str], list[list]]:
    """读取工作表 → (表头, 数据行)。诊断信息见 read_sheet_meta()"""
    headers, rows, _ = read_sheet_meta(path, sheet)
    return headers, rows


def make_norm(opts: dict):
    """按选项生成单元格归一化函数（与 webapp 的比对口径保持一致）"""
    def norm(v):
        if v is None:
            return "" if opts.get("nullify", True) else "\x00NULL"
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        s = str(v)
        if opts.get("strip", True):
            s = s.strip()
        if not opts.get("strict_num", False):
            try:
                s = f"{float(s):.10g}"
            except ValueError:
                pass
        if not opts.get("case", False):
            s = s.lower()
        return s
    return norm


def compare_excel(pa: Path, pb: Path, keys: list[str], opts: dict) -> dict:
    """按主键对齐比对两个 Excel，返回统计与差异明细"""
    ha, ra = read_sheet(pa)
    hb, rb = read_sheet(pb)
    norm = make_norm(opts)
    ia = {h: i for i, h in enumerate(ha)}
    ib = {h: i for i, h in enumerate(hb)}
    missing = [k for k in keys if k not in ia or k not in ib]
    if missing:
        raise RuntimeError(f"主键列在两个文件中不都存在: {missing}")
    cmp_cols = [c for c in ha if c in ib and c not in keys]

    def build(rows, idx):
        """键用归一化值做匹配，同时留一份原始值供展示，避免界面显示被 lower 过的键"""
        m = {}
        for r in rows:
            k = tuple(norm(r[idx[c]]) if idx[c] < len(r) else "" for c in keys)
            raw = tuple("" if idx[c] >= len(r) or r[idx[c]] is None
                        else str(r[idx[c]]) for c in keys)
            m.setdefault(k, []).append((r, raw))
        return m

    ma, mb = build(ra, ia), build(rb, ib)
    only_a = [ma[k][0][1] for k in ma if k not in mb]
    only_b = [mb[k][0][1] for k in mb if k not in ma]
    matched, diffs = 0, []
    for k in ma:
        if k not in mb:
            continue
        for (rowa, raw), (rowb, _) in zip(ma[k], mb[k]):
            matched += 1
            for c in cmp_cols:
                va = rowa[ia[c]] if ia[c] < len(rowa) else None
                vb = rowb[ib[c]] if ib[c] < len(rowb) else None
                if norm(va) != norm(vb):
                    diffs.append({"key": "+".join(raw), "col": c,
                                  "a": "" if va is None else str(va),
                                  "b": "" if vb is None else str(vb)})
    return {"stats": {"rows_a": len(ra), "rows_b": len(rb),
                      "only_a": len(only_a), "only_b": len(only_b),
                      "matched": matched, "cols": len(cmp_cols),
                      "diff_cells": len(diffs)},
            "only_a": ["+".join(k) for k in only_a],
            "only_b": ["+".join(k) for k in only_b],
            "diffs": diffs, "keys": keys}


# ── 写出：Excel / CSV ───────────────────────────────────────────

_HEADER_FILL = "2F5496"


def write_xlsx(path: Path, sheets: dict[str, tuple[list[str], list[list]]]) -> Path:
    """写多 sheet Excel。sheets = {页名: (表头, 行)}，带表头样式与列宽自适应"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    for title, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title[:31])
        if headers:
            ws.append(headers)
            fill = PatternFill("solid", fgColor=_HEADER_FILL)
            font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for c in ws[1]:
                c.fill, c.font, c.alignment = fill, font, center
            ws.freeze_panes = "A2"
        for r in rows:
            ws.append(["" if v is None else v for v in r])
        # 列宽：按内容估算，中文字符按 2 个宽度计
        ncol = len(headers) if headers else (max((len(r) for r in rows), default=0))
        for i in range(1, ncol + 1):
            vals = [headers[i - 1]] if headers and i <= len(headers) else []
            vals += [str(r[i - 1]) for r in rows[:300] if i <= len(r) and r[i - 1] is not None]
            w = max((len(v) + sum(1 for ch in v if ord(ch) > 127) for v in vals), default=8)
            ws.column_dimensions[get_column_letter(i)].width = min(max(w + 3, 9), 52)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_csv(path: Path, headers: list[str], rows: list[list]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        if headers:
            w.writerow(headers)
        for r in rows:
            w.writerow(["" if v is None else v for v in r])
    return path


def safe_name(name: str) -> str:
    """把用户输入的文件名清成安全字符，防止路径穿越与非法字符"""
    cleaned = re.sub(r"[^\w\u4e00-\u9fff\-]", "_", str(name).strip())
    return (cleaned[:80] or "导出")


def detect_delimiter(line: str) -> str:
    """自动识别分隔符：Tab > 竖线 > 逗号 > 连续空格"""
    for d in ("\t", "|", ","):
        if d in line:
            return d
    return "  "


def split_line(line: str, delim: str) -> list[str]:
    import re
    if delim == "  ":
        return [c for c in re.split(r"\s{2,}", line.strip()) if c]
    return [c.strip() for c in line.split(delim)]


def text_to_rows(text: str, delim: str = "auto") -> list[list[str]]:
    """粘贴的表格文本 → 补齐列数的二维数组"""
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        raise RuntimeError("没有可解析的文本")
    if delim == "auto":
        delim = detect_delimiter(lines[0])
    rows = [split_line(l, delim) for l in lines]
    n = max(len(r) for r in rows)
    return [r + [""] * (n - len(r)) for r in rows]
